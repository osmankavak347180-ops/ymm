"""Mizan Excel parser testleri (mizan_oku)."""

from decimal import Decimal
from pathlib import Path
import sys
import importlib.util

import pytest
import yaml
from openpyxl import Workbook

from ymm.parsers.mizan import mizan_oku

PROJE_KOKU = Path(__file__).resolve().parent.parent
DUMMY_MIZAN = PROJE_KOKU / "ornek_veri" / "mizan_2025.xlsx"
HARITA_YOLU = PROJE_KOKU / "config" / "kolon_haritasi.yaml"


def _varsayilan_harita() -> dict:
    with open(HARITA_YOLU, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _uret_modulunu_import_et():
    """ornek_veri.uret modülünü importlib ile yükle (paket değil)."""
    uret_yolu = PROJE_KOKU / "ornek_veri" / "uret.py"
    spec = importlib.util.spec_from_file_location("ornek_veri_uret", uret_yolu)
    uret_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(uret_mod)
    return uret_mod


def test_dummy_mizan_satir_sayisi_ve_600_ankor_deger():
    """Dummy mizanda 28 satir olmali; 600 hesabinin alacak_bakiye'si tam 5.000.000,00 olmali."""
    harita = _varsayilan_harita()

    satirlar = mizan_oku(DUMMY_MIZAN, harita)

    assert len(satirlar) == 28

    hesap_600 = next(s for s in satirlar if s.hesap_kodu == "600")
    assert hesap_600.alacak_bakiye == Decimal("5000000.00")
    assert isinstance(hesap_600.alacak_bakiye, Decimal)


def test_dummy_mizan_diger_ankor_degerler():
    """Faz 1-2 testlerinin dayanacagi diger sabit degerler: 770, 131, 689."""
    harita = _varsayilan_harita()
    satirlar = mizan_oku(DUMMY_MIZAN, harita)
    satir_map = {s.hesap_kodu: s for s in satirlar}

    assert satir_map["770"].borc_bakiye == Decimal("800000.00")
    assert satir_map["131"].borc_bakiye == Decimal("150000.00")
    assert satir_map["689"].borc_bakiye == Decimal("45000.00")


def test_dummy_mizan_borc_alacak_dengede():
    """Uretilen dummy mizanin borc toplam/bakiye toplamlari alacak ile esit olmali."""
    harita = _varsayilan_harita()
    satirlar = mizan_oku(DUMMY_MIZAN, harita)

    toplam_borc = sum((s.borc_toplam for s in satirlar), start=Decimal("0"))
    toplam_alacak = sum((s.alacak_toplam for s in satirlar), start=Decimal("0"))
    toplam_borc_bakiye = sum((s.borc_bakiye for s in satirlar), start=Decimal("0"))
    toplam_alacak_bakiye = sum((s.alacak_bakiye for s in satirlar), start=Decimal("0"))

    assert toplam_borc == toplam_alacak
    assert toplam_borc_bakiye == toplam_alacak_bakiye


def test_bos_satir_atlanir(tmp_path):
    """hesap_kodu hucresi bos olan satirlar sonuca dahil edilmez."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Hesap Kodu", "Hesap Adı", "Borç Toplam", "Alacak Toplam", "Borç Bakiye", "Alacak Bakiye"])
    ws.append(["100", "Kasa", 1000.0, 0.0, 1000.0, 0.0])
    ws.append([None, None, None, None, None, None])  # bos satir
    ws.append(["102", "Bankalar", 2000.0, 0.0, 2000.0, 0.0])
    dosya = tmp_path / "test_bos_satir.xlsx"
    wb.save(dosya)

    harita = {
        "hesap_kodu": "A",
        "hesap_adi": "B",
        "borc_toplam": "C",
        "alacak_toplam": "D",
        "borc_bakiye": "E",
        "alacak_bakiye": "F",
    }

    satirlar = mizan_oku(dosya, harita)

    assert len(satirlar) == 2
    assert [s.hesap_kodu for s in satirlar] == ["100", "102"]


def test_turkce_bicimli_tutar_normalize_edilir(tmp_path):
    """'1.234.567,89' bicimindeki string hucre Decimal('1234567.89') olarak okunmali."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Hesap Kodu", "Hesap Adı", "Borç Toplam", "Alacak Toplam", "Borç Bakiye", "Alacak Bakiye"])
    ws.append(["120", "Alıcılar", "1.234.567,89", "0,00", "1.234.567,89", "0,00"])
    dosya = tmp_path / "test_turkce_bicim.xlsx"
    wb.save(dosya)

    harita = {
        "hesap_kodu": "A",
        "hesap_adi": "B",
        "borc_toplam": "C",
        "alacak_toplam": "D",
        "borc_bakiye": "E",
        "alacak_bakiye": "F",
    }

    satirlar = mizan_oku(dosya, harita)

    assert len(satirlar) == 1
    assert satirlar[0].borc_toplam == Decimal("1234567.89")
    assert isinstance(satirlar[0].borc_toplam, Decimal)


def test_harita_baslik_adiyla_calisir(tmp_path):
    """Harita degerleri kolon harfi yerine baslik metni de olabilir (esnek parser)."""
    wb = Workbook()
    ws = wb.active
    # Kolon sirasi degisik / farkli baslik adlari (mukellefe ozel format simulasyonu)
    ws.append(["Kod", "Ad", "Bakiye Borc", "Bakiye Alacak", "Toplam Borc", "Toplam Alacak"])
    ws.append(["131", "Ortaklardan Alacaklar", 150000.0, 0.0, 150000.0, 0.0])
    dosya = tmp_path / "test_baslik_haritasi.xlsx"
    wb.save(dosya)

    harita = {
        "hesap_kodu": "Kod",
        "hesap_adi": "Ad",
        "borc_bakiye": "Bakiye Borc",
        "alacak_bakiye": "Bakiye Alacak",
        "borc_toplam": "Toplam Borc",
        "alacak_toplam": "Toplam Alacak",
    }

    satirlar = mizan_oku(dosya, harita)

    assert len(satirlar) == 1
    assert satirlar[0].hesap_kodu == "131"
    assert satirlar[0].borc_bakiye == Decimal("150000.00")


def test_bos_dosya_sadece_baslik_bos_liste_doner(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Hesap Kodu", "Hesap Adı", "Borç Toplam", "Alacak Toplam", "Borç Bakiye", "Alacak Bakiye"])
    dosya = tmp_path / "test_bos_dosya.xlsx"
    wb.save(dosya)

    harita = {
        "hesap_kodu": "A",
        "hesap_adi": "B",
        "borc_toplam": "C",
        "alacak_toplam": "D",
        "borc_bakiye": "E",
        "alacak_bakiye": "F",
    }

    assert mizan_oku(dosya, harita) == []


def test_uret_modulu_ile_olusturulan_mizan_ankor_degerler(tmp_path):
    """uret() ile dinamik oluşturulan mizan dosyasının ankor değerlerini doğrula."""
    uret_mod = _uret_modulunu_import_et()
    uret = uret_mod.uret

    # uret() çalıştır, tmp_path'a çıktı yaz
    dosya = tmp_path / "m.xlsx"
    uret(dosya)

    # Oluşturulan dosyayı oku
    harita = _varsayilan_harita()
    satirlar = mizan_oku(dosya, harita)
    satir_map = {s.hesap_kodu: s for s in satirlar}

    # Ankor değerleri assert et
    assert satir_map["600"].alacak_bakiye == Decimal("5000000.00")
    assert satir_map["770"].borc_bakiye == Decimal("800000.00")
    assert satir_map["131"].borc_bakiye == Decimal("150000.00")
    assert satir_map["689"].borc_bakiye == Decimal("45000.00")


def test_uret_modulu_ile_olusturulan_mizan_denge(tmp_path):
    """uret() ile oluşturulan mizan dosyasında borç ve alacak dengesi doğru mu."""
    uret_mod = _uret_modulunu_import_et()
    uret = uret_mod.uret

    dosya = tmp_path / "m.xlsx"
    uret(dosya)

    harita = _varsayilan_harita()
    satirlar = mizan_oku(dosya, harita)

    # Denge kontrol
    toplam_borc = sum((s.borc_toplam for s in satirlar), start=Decimal("0"))
    toplam_alacak = sum((s.alacak_toplam for s in satirlar), start=Decimal("0"))
    toplam_borc_bakiye = sum((s.borc_bakiye for s in satirlar), start=Decimal("0"))
    toplam_alacak_bakiye = sum((s.alacak_bakiye for s in satirlar), start=Decimal("0"))

    assert toplam_borc == toplam_alacak
    assert toplam_borc_bakiye == toplam_alacak_bakiye


def test_bos_hesap_kodu_boş_string_atlanir(tmp_path):
    """hesap_kodu hucresi bos string ('') olan satirlar sonuca dahil edilmez."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Hesap Kodu", "Hesap Adı", "Borç Toplam", "Alacak Toplam", "Borç Bakiye", "Alacak Bakiye"])
    ws.append(["100", "Kasa", 1000.0, 0.0, 1000.0, 0.0])
    ws.append(["", "Boş Hesap", 500.0, 0.0, 500.0, 0.0])  # boş string hesap_kodu
    ws.append(["102", "Bankalar", 2000.0, 0.0, 2000.0, 0.0])
    dosya = tmp_path / "test_bos_hesap_kodu.xlsx"
    wb.save(dosya)

    harita = {
        "hesap_kodu": "A",
        "hesap_adi": "B",
        "borc_toplam": "C",
        "alacak_toplam": "D",
        "borc_bakiye": "E",
        "alacak_bakiye": "F",
    }

    satirlar = mizan_oku(dosya, harita)

    assert len(satirlar) == 2
    assert [s.hesap_kodu for s in satirlar] == ["100", "102"]
