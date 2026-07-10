"""Mizan (kesin/genel gecici mizan) Excel parser.

`harita` (kolon_haritasi.yaml'dan yuklenen dict) her `MizanSatiri` alani
icin bir "kolon adresi" verir:
    {"hesap_kodu": "A", "hesap_adi": "B", "borc_toplam": "C", ...}

Adres ya Excel kolon harfidir ("A", "AB" -- yalnizca A-Z, 1-3 karakter,
tamami buyuk harf) ya da 1. satirdaki baslik hucresinin metnidir (orn.
"Hesap Kodu"). Boylece mukellefe gore degisen mizan formatlarina
(farkli kolon sirasi / baslik adi) kolon_haritasi.yaml duzenlenerek
uyum saglanir (bkz. docs/01-MIMARI.md, R8).
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from ymm.modeller import MizanSatiri

_KOLON_HARFI_RE = re.compile(r"^[A-Z]{1,3}$")

# MizanSatiri alan sirasi (uretilen liste bu sirayla doldurulur).
_ALAN_SIRASI = (
    "hesap_kodu",
    "hesap_adi",
    "borc_toplam",
    "alacak_toplam",
    "borc_bakiye",
    "alacak_bakiye",
)

_TUTAR_ALANLARI = ("borc_toplam", "alacak_toplam", "borc_bakiye", "alacak_bakiye")


def _tutar_normalize(deger: object) -> Decimal:
    """Hucre degerini Decimal'e cevirir.

    - Bos (None) veya bos string -> Decimal("0")
    - Sayisal hucre (int/float) -> `Decimal(str(deger))` uzerinden (dogrudan
      `Decimal(float)` YAPILMAZ; hassasiyet bozulmasin diye).
    - String hucre -> Turk bicimi normalize edilir:
      "1.234.567,89" -> "1234567.89" (binlik "." kaldirilir, ondalik "," -> ".")
    """
    if deger is None:
        return Decimal("0")
    if isinstance(deger, Decimal):
        return deger
    if isinstance(deger, bool):
        raise ValueError(f"Gecersiz tutar hucresi (bool): {deger!r}")
    if isinstance(deger, (int, float)):
        return Decimal(str(deger))
    if isinstance(deger, str):
        metin = deger.strip()
        if not metin:
            return Decimal("0")
        metin = metin.replace(".", "").replace(",", ".")
        try:
            return Decimal(metin)
        except InvalidOperation as exc:
            raise ValueError(f"Gecersiz tutar degeri: {deger!r}") from exc
    raise ValueError(f"Desteklenmeyen hucre tipi: {type(deger)!r} ({deger!r})")


def _kolon_index_bul(harita_degeri: str, basliklar: dict[str, int]) -> int:
    """Harita degerini (kolon harfi ya da baslik metni) 1-index kolon numarasina cozer."""
    if _KOLON_HARFI_RE.match(harita_degeri):
        return column_index_from_string(harita_degeri)
    try:
        return basliklar[harita_degeri]
    except KeyError as exc:
        raise KeyError(
            f"Kolon haritasi degeri {harita_degeri!r}: ne kolon harfi ne de "
            f"1. satirdaki basliklar arasinda bulunan bir baslik. "
            f"Mevcut basliklar: {list(basliklar)}"
        ) from exc


def _satirlari_oku(dosya: Path) -> list[tuple]:
    """Dosya uzantisina gore tum satirlari tuple listesi olarak dondurur.

    - .xlsx / .xlsm: openpyxl (values_only)
    - .xls (Excel 97-2003): xlrd. xlrd sayisal hucreyi daima float verir;
      tam sayilar int'e cevrilir ki hesap kodlari "100.0" degil "100" okunssun
      (openpyxl davranisiyla ayni).
    """
    if dosya.suffix.lower() == ".xls":
        import xlrd

        kitap = xlrd.open_workbook(str(dosya))
        sayfa = kitap.sheet_by_index(0)

        def _deger(hucre):
            if hucre.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                return None
            v = hucre.value
            if isinstance(v, float) and v.is_integer():
                return int(v)
            return v

        return [
            tuple(_deger(h) for h in sayfa.row(i)) for i in range(sayfa.nrows)
        ]

    wb = load_workbook(dosya, data_only=True)
    ws = wb.active
    return [tuple(r) for r in ws.iter_rows(values_only=True)]


def mizan_oku(dosya: Path, harita: dict) -> list[MizanSatiri]:
    """Excel mizan dosyasini (xlsx/xlsm/xls) haritaya gore okur.

    - 1. satir baslik satiri kabul edilir (kolon adi olarak kullanilir).
    - `hesap_kodu` hucresi bos olan satirlar atlanir (bos satir).
    - Tutar alanlari Decimal'e normalize edilir.
    """
    tum_satirlar = _satirlari_oku(dosya)
    if not tum_satirlar:
        return []

    ilk_satir = tum_satirlar[0]
    basliklar = {
        str(deger).strip(): idx + 1
        for idx, deger in enumerate(ilk_satir)
        if deger is not None
    }

    kolon_index = {alan: _kolon_index_bul(harita[alan], basliklar) for alan in _ALAN_SIRASI}

    satirlar: list[MizanSatiri] = []
    for row in tum_satirlar[1:]:
        hesap_kodu_ham = row[kolon_index["hesap_kodu"] - 1]
        if hesap_kodu_ham is None or str(hesap_kodu_ham).strip() == "":
            continue  # bos satir atla

        alanlar: dict[str, object] = {
            "hesap_kodu": str(hesap_kodu_ham).strip(),
            "hesap_adi": str(row[kolon_index["hesap_adi"] - 1] or "").strip(),
        }
        for alan in _TUTAR_ALANLARI:
            alanlar[alan] = _tutar_normalize(row[kolon_index[alan] - 1])

        satirlar.append(MizanSatiri(**alanlar))

    return satirlar
